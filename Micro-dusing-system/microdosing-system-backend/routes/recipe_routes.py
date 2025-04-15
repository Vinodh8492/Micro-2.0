from flask import Blueprint, request, jsonify # type: ignore
from extensions import db
from models.recipe import Recipe, RecipeMaterial
from models.production import ProductionOrder
from models.user import User
from sqlalchemy.exc import IntegrityError
from flask import send_file, jsonify
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image as PILImage
import io, os, tempfile


recipe_bp = Blueprint("recipe", __name__)

@recipe_bp.route("/recipes/export/barcodes", methods=["GET"])
def export_recipes_excel_with_barcodes():
    try:
        recipes = Recipe.query.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Recipes with Barcodes"
        ws.append(["Name", "Code", "Barcode ID", "Scannable Barcode"])

        row_number = 2

        for recipe in recipes:
            if recipe.barcode_id:
                barcode_id = recipe.barcode_id
                try:
                    # Create barcode image in temp dir
                    temp_dir = tempfile.gettempdir()
                    filename = f"{barcode_id}"
                    filepath = os.path.join(temp_dir, f"{filename}.png")

                    code128 = Code128(barcode_id, writer=ImageWriter())
                    code128.save(os.path.join(temp_dir, filename))  # Note: no .png in save()

                    # Resize for Excel
                    image = PILImage.open(filepath)
                    image = image.resize((200, 60))
                    image.save(filepath)

                    # Add data to Excel
                    ws.cell(row=row_number, column=1, value=recipe.name)
                    ws.cell(row=row_number, column=2, value=recipe.code)
                    ws.cell(row=row_number, column=3, value=barcode_id)

                    img = ExcelImage(filepath)
                    img.width = 150
                    img.height = 50
                    ws.add_image(img, f"D{row_number}")

                    os.remove(filepath)

                    row_number += 1

                except Exception as e:
                    print(f"Failed to generate barcode for {barcode_id}: {e}")
                    ws.cell(row=row_number, column=1, value=recipe.name)
                    ws.cell(row=row_number, column=2, value=recipe.code)
                    ws.cell(row=row_number, column=3, value=barcode_id)
                    row_number += 1

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return send_file(
            stream,
            download_name="recipes_with_barcodes.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@recipe_bp.route("/recipes", methods=["POST"])
def create_recipe():
    data = request.get_json()

    required_fields = ["name", "code", "version", "created_by"]
    for field in required_fields:
        if not data.get(field):
            return jsonify({"error": f"'{field}' is required."}), 400

    # ✅ Validate status
    status = data.get("status", "Unreleased")
    valid_statuses = ["Released", "Unreleased"]
    if status not in valid_statuses:
        return jsonify({"error": f"Invalid status value: {status}"}), 400

    # ✅ Validate user existence
    user = db.session.get(User, data["created_by"])
    if not user:
        return jsonify({"error": "User not found."}), 400

    # Validate no_of_materials
    no_of_materials = data.get("no_of_materials")
    if no_of_materials is not None:
        try:
            no_of_materials = int(no_of_materials)
            if no_of_materials < 0:
                raise ValueError
        except (ValueError, TypeError):
            return jsonify({"error": "no_of_materials must be a non-negative integer."}), 400

        

    new_recipe = Recipe(
        name=data["name"],
        code=data["code"],
        description=data.get("description"),
        version=data["version"],
        status=status,
        created_by=data["created_by"],
        barcode_id=data.get("barcode_id"),
        no_of_materials=no_of_materials
    )

    db.session.add(new_recipe)

    try:
        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        if "Duplicate entry" in str(e.orig):
            return jsonify({"error": "Duplicate entry: code or barcode_id already exists."}), 400
        return jsonify({"error": "Database error occurred."}), 500
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

    return jsonify({"message": "Recipe created successfully!"}), 201

@recipe_bp.route("/recipes/<int:recipe_id>", methods=["GET"])
def get_recipe(recipe_id):
    recipe = Recipe.query.get(recipe_id)
    if not recipe:
        return jsonify({"error": "Recipe not found"}), 404

    result = {
        "recipe_id": recipe.recipe_id,
        "name": recipe.name,
        "code": recipe.code,
        "description": recipe.description,
        "version": recipe.version,
        "status": recipe.status,
        "created_by": recipe.created_by,
    }
    return jsonify(result)

@recipe_bp.route("/recipes", methods=["GET"])
def get_recipes():
    recipes = Recipe.query.all()
    result = [
        {
            "recipe_id": recipe.recipe_id,
            "name": recipe.name,
            "code": recipe.code,
            "description": recipe.description,
            "version": recipe.version,
            "status": recipe.status,
            "created_by": recipe.created_by,
            "created_at" : recipe.created_at,
            "no_of_materials" : recipe.no_of_materials,
        }
        for recipe in recipes
    ]
    return jsonify(result)

@recipe_bp.route("/recipes/<int:recipe_id>", methods=["PUT"])
def update_recipe(recipe_id):
    recipe = Recipe.query.get(recipe_id)
    if not recipe:
        return jsonify({"message": "Recipe not found"}), 404

    data = request.get_json()
    recipe.name = data.get("name", recipe.name)
    recipe.code = data.get("code", recipe.code)
    recipe.description = data.get("description", recipe.description)
    recipe.version = data.get("version", recipe.version)
    recipe.status = data.get("status", recipe.status)
    recipe.no_of_materials = data.get("no_of_materials", recipe.no_of_materials)

    db.session.commit()
    return jsonify({"message": "Recipe updated successfully"}), 200
@recipe_bp.route("/recipes/<int:recipe_id>", methods=["DELETE"])
def delete_recipe(recipe_id):
    try:
        # Step 1: Delete related records in `production_order`
        # Here, you can also use cascade delete by modifying the ForeignKey in your ProductionOrder model
        db.session.query(ProductionOrder).filter(ProductionOrder.recipe_id == recipe_id).delete(synchronize_session=False)

        # Step 2: Update `recipe_id` to NULL for related `recipe_material` records
        # Using `synchronize_session=False` to avoid session issues with large batches
        db.session.query(RecipeMaterial).filter(RecipeMaterial.recipe_id == recipe_id).update(
            {RecipeMaterial.recipe_id: None}, synchronize_session=False
        )

        # Step 3: Now delete the recipe itself
        recipe = Recipe.query.get(recipe_id)
        if not recipe:
            return jsonify({"message": "Recipe not found"}), 404

        db.session.delete(recipe)
        db.session.commit()  # Commit the deletions
        
        return jsonify({"message": "Recipe deleted successfully"}), 200

    except Exception as e:
        db.session.rollback()  # Rollback in case of an error
        app.logger.error(f"Error deleting recipe: {str(e)}")
        return jsonify({"error": str(e)}), 500


### 🚀 RECIPE MATERIAL ROUTES ###
@recipe_bp.route("/recipe_materials", methods=["POST"])
def create_recipe_material():
    data = request.get_json()
    new_recipe_material = RecipeMaterial(
        recipe_id=data["recipe_id"],
        material_id=data["material_id"],
        quantity=data["quantity"],
        sequence_number=data["sequence_number"],
    )
    db.session.add(new_recipe_material)
    db.session.commit()
    return jsonify({"message": "Recipe material added successfully!"}), 201

@recipe_bp.route("/recipe_materials/<int:recipe_id>", methods=["GET"])
def get_recipe_materials(recipe_id):
    materials = RecipeMaterial.query.filter_by(recipe_id=recipe_id).all()
    result = [
        {
            "recipe_material_id": mat.recipe_material_id,
            "recipe_id": mat.recipe_id,
            "material_id": mat.material_id,
            "quantity": str(mat.quantity),
            "sequence_number": mat.sequence_number,
        }
        for mat in materials
    ]
    return jsonify(result)

@recipe_bp.route("/recipe_materials/<int:recipe_material_id>", methods=["PUT"])
def update_recipe_material(recipe_material_id):
    material = RecipeMaterial.query.get(recipe_material_id)
    if not material:
        return jsonify({"message": "Recipe material not found"}), 404

    data = request.get_json()
    material.quantity = data.get("quantity", material.quantity)
    material.sequence_number = data.get("sequence_number", material.sequence_number)

    db.session.commit()
    return jsonify({"message": "Recipe material updated successfully"}), 200

@recipe_bp.route("/recipe_materials/<int:recipe_material_id>", methods=["DELETE"])
def delete_recipe_material(recipe_material_id):
    material = RecipeMaterial.query.get(recipe_material_id)
    if not material:
        return jsonify({"message": "Recipe material not found"}), 404

    db.session.delete(material)
    db.session.commit()
    return jsonify({"message": "Recipe material deleted successfully"}), 200
